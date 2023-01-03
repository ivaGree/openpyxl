# less_14.py - Посчитаем сумму по группировке поля

# Подключили openpyxl
import openpyxl
from openpyxl import *
from itertools import groupby

# Подключаем эксельку с нашей базой
wb = load_workbook("database.xlsx")
ws = wb['worklog']

# Создадим пустой список списков
staff_row_list: list = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    row = [cell.value for cell in row]
    # print(row)  # ['ID', 'Фио', 'Логин', 'Роль'] ... и далее
    # Добавляем данные в список списков
    staff_row_list.append(row)

print(staff_row_list)

# Пример ('Категория', 'Товар', 'Цена', 'Номер склада') / Надо сгруппировать списки по категории и номеру склада, так чтобы цены на каждый товар из категории на определенном складе тоже сложились.
# list_in = [['Игрушки', 'Мяч надувной', '300', '1'], ['Игрушки', 'Бионикл', '8000', '1'], ['Ткань', 'Вельвет', '1000', '2'], ['Ткань', 'Джинса', '500', '2'], ['Игрушки', 'Бионикл', '1000', '2']]
# list_out = [[*k, sum(int(rr[2]) for rr in v)] for k, v in groupby(list_in, lambda r: (r[0], r[-1]))]
# print(list_out)

staff_row_list_out = [[*k, sum(float(rr[1]) for rr in v)] for k, v in groupby(staff_row_list, lambda r: (r[3],))]
print(staff_row_list_out)