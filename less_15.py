# less_15.py - Применим (ВПР) в рамках двух листов

import openpyxl
from openpyxl import load_workbook
import datetime as dt

# Подключаем эксельку с нашей базой
wb = load_workbook("database.xlsx")
ws_staff = wb['staff']
ws_worklog = wb['worklog']

# i = 1
# for staff_login in ws_staff.values:
#     workTimeSum = 0
#     login = ws_staff[i][2].value
#     print(login)
#
#     for worklog_login in ws_worklog.values:
#         # if login == worklog_login[3]:
#         #     workTimeSum += login
#         # print(workTimeSum)
#         parsedUsers = (ws_staff[i][2].value == worklog_login[3])
#     print(parsedUsers)
#     print('---')
#     i += 1

for row in ws_staff.iter_rows(min_row=2, max_row=ws_staff.max_row, min_col=1, max_col=ws_staff.max_column):
    row = [cell.value for cell in row]
    login = row[2]
    # print(f'{login} ---')

    sum = 0
    for row in ws_worklog.iter_rows(min_row=2, max_row=ws_worklog.max_row, min_col=1, max_col=ws_worklog.max_column):
        row = [cell.value for cell in row]
        login2 = row[3]
        worklog = row[1]
        if login == login2:
            # print(f'\t{login2} + {worklog}')
            sum += worklog
    # print(sum)
    print(f'{login} --- {sum}')





