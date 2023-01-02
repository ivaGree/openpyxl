# less7.py - Вариант 1 (for row in ws.iter_rows()): Читаем лист Excel файла (предпочтительно) -> в список списков

# Подключили openpyxl
import openpyxl
from openpyxl import *

# Подключаем эксельку с нашей базой
wb = load_workbook("database.xlsx")

# Постучаться к листу (staff) можно одним из 3х способов / <Worksheet "staff">
# ws = wb.active        # 1) по активному листу
# ws = wb.worksheets[0] # 2) по порядковому номеру в книге, если в файле поменять местами листы - то все поломается
ws = wb['staff']        # 3) по конкретному названию листа в книге

# Создадим пустой список списков
staff_row_list: list = []

# ПРЕДПОЧТИТЕЛЬНЫЙ МЕТОД ДЛЯ ПЕРЕБОРА ТАБЛИЦ
# min_row=1 -> прочитает с заголовком
# min_row=2 -> без заголовка начиная со 2ой строки
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    # print(row) # (<Cell 'staff'.A1>, <Cell 'staff'.B1>, <Cell 'staff'.C1>, <Cell 'staff'.D1>) ... и далее
    row = [cell.value for cell in row]
    # print(row)  # ['ID', 'Фио', 'Логин', 'Роль'] ... и далее
    # Добавляем данные в список списков
    staff_row_list.append(row)

print(staff_row_list)   # [['ID', 'Фио', 'Логин', 'Роль'], [11, 'Васечкин Петя', 'vasechkin', 'Тестировщик'] ... [и далее]]