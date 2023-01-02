# less_13.py - Прочитаем database.xlsx лист worklog и запишем выборочные данные в tmp13.xlsx (+ переименуем заголовки + удалим лишние пробелы в ячейках + обработаем дату)

import openpyxl
from openpyxl import load_workbook
import datetime as dt

# Подключаем эксельку с нашей базой
wb_database = load_workbook("database.xlsx")    # book_base
# Постучаться к листу (staff)
ws_worklog = wb_database['worklog']                 # worklog_tempo

# ----------------------------------------------------------------------------------------------------------------------
# Прочитаем файлик (database.xlsx) и загоним выборочные данные листа (staff) в список списков (ws_staff_list)
# Создали пустой список списков
ws_worklog_list: list = []

# Перебираем файл без 1 строки без заголовков -> только данные
for row in ws_worklog.iter_rows(min_row=2, max_row=ws_worklog.max_row, min_col=1, max_col=ws_worklog.max_column):
    row = [cell.value for cell in row]
    date_worklog = row[2].strftime("%d.%m.%Y")      # Получили все данные 3го столбца C: Дата списания / преобразуем в формат 01.01.2023
    hours = row[1]             # Получили все данные 2го столбца B: Списано часов
    login = row[3].strip()     # Получили все данные 4го столбца D: Логин пользователя / почистили пробелы в начале и в конце

    # Запишем все данные выбранных столбцов во временный список
    tmp_list = [date_worklog, hours, login]
    # Загоним в список списков выбранные значения листа временного списка
    ws_worklog_list.append(tmp_list)

print(ws_worklog_list)
# ----------------------------------------------------------------------------------------------------------------------
# Запишем в файлик
# Тестовая таблица куда будем писать данные и которая будет создаваться сама
wb_test = openpyxl.Workbook()
# по умолчанию создается книга с листом Sheet, снесем его
wb_test.remove(wb_test.active)
# Создаем лист с названием "Test"
wb_test.create_sheet('Test')
# Делаем его активным
ws_test = wb_test['Test']

# Бежим по списку списков ws_staff_list
for row in ws_worklog_list:
    ws_test.append(row)

# Вставляем столбец с заголовками
ws_test.insert_rows(0)
ws_test['A1'].value = 'date_worklog'
ws_test['B1'].value = 'hours'
ws_test['C1'].value = 'login'

# Сохраняем измененный файл
wb_test.save('tmp13.xlsx')