# Less4: Определяем количество строк и столбцов на листе со значениями

import openpyxl
from openpyxl import load_workbook

# ----------------------------------------------------------------------------------------------------------------------
# Подключаем эксельку с нашей базой
wb = load_workbook("database.xlsx")

# Записываем листы в переменные
ws = wb['staff']          # ws -> ws_staff

# row       строки
# column    столбцы
# cell      ячейка
# Определяем количество строк на листе / staff
row_count = ws.max_row
print(f'Строк: {row_count}')
# Определяем количество столбцов на листе / staff
column_count = ws.max_column
print(f'Столбцов: {column_count}')