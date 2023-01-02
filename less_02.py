# Less2: Полный перечень всех листов книги в виде списка

import openpyxl
from openpyxl import load_workbook

# ----------------------------------------------------------------------------------------------------------------------
# Подключаем эксельку с нашей базой
wb = load_workbook("database.xlsx")

# Запишем полный перечень всех листов книги в список
wb_sheets_list = wb.sheetnames
print(wb_sheets_list)                         # ['staff', 'worklog']

print(wb_sheets_list[0])                      # Обратиться к 1му листу книги -> staff
print(type(wb_sheets_list))                   # ТИП: <class 'list'>
print(type(wb_sheets_list[0]))                # ТИП: <class 'str'>