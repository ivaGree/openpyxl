# less_12.py - Прочитаем database.xlsx и запишем выборочные данные в tmp12.xlsx (+ переименуем заголовки)

import openpyxl
from openpyxl import load_workbook
import datetime as dt

# Подключаем эксельку с нашей базой
wb_database = load_workbook("database.xlsx")    # book_base
# Постучаться к листу (staff)
ws_staff = wb_database['staff']                 # worklog_tempo

# ----------------------------------------------------------------------------------------------------------------------
# Прочитаем файлик (database.xlsx) и загоним выборочные данные листа (staff) в список списков (ws_staff_list)
# Создали пустой список списков
ws_staff_list: list = []

# Перебираем файл без 1 строки без заголовков -> только данные
for row in ws_staff.iter_rows(min_row=2, max_row=ws_staff.max_row, min_col=1, max_col=ws_staff.max_column):
    row = [cell.value for cell in row]
    user_name = row[1]      # Получили все данные 2го столбца B: Фио
    login = row[2]          # Получили все данные 3го столбца C: Логин

    # Запишем все данные выбранных столбцов во временный список
    tmp_list = [user_name, login]
    # Загоним в список списков выбранные значения листа временного списка
    ws_staff_list.append(tmp_list)

print(ws_staff_list)
# ----------------------------------------------------------------------------------------------------------------------
# Запишем в файлик
# Тестовая таблица куда будем писать данные и которая будет создаваться сама
wb_test = openpyxl.Workbook()
# по умолчанию создается книга с листом Sheet, снесем его
wb_test.remove(wb_test.active)
# Создаем лист с названием "Test"
wb_test.create_sheet('Test')
# Делаем его активным
ws_test = wb_test['Test']

# Бежим по списку списков ws_staff_list
for row in ws_staff_list:
    ws_test.append(row)

# Вставляем столбец с заголовками
ws_test.insert_rows(0)
ws_test['A1'].value = 'user_name'
ws_test['B1'].value = 'login'

# Сохраняем измененный файл
wb_test.save('tmp12.xlsx')
