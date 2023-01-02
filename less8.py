# less8.py - Вариант 2 (for row in range(1, ws.max_row+1)): Читаем лист Excel файла -> в список списков

# Подключили openpyxl
import openpyxl
from openpyxl import *

# Подключаем эксельку с нашей базой
wb = load_workbook("database.xlsx")

# Постучаться к листу (staff) можно одним из 3х способов / <Worksheet "staff">
# ws = wb.active        # 1) по активному листу
# ws = wb.worksheets[0] # 2) по порядковому номеру в книге, если в файле поменять местами листы - то все поломается
ws = wb['staff']        # 3) по конкретному названию листа в книге

# Создадим пустой список
staff_list = []
# Создадим пустой список списков
staff_row_list: list = []

# (1, ws_staff.max_row+1) -> с заголовком, прочитает весь лист
# (2, ws_staff.max_row+1) -> без заголовка, начнет со второй строки
for row in range(1, ws.max_row+1):
    # print(ws_staff[row])    # вывод: (<Cell 'staff'.A1>, <Cell 'staff'.B1>, <Cell 'staff'.C1>)
    staff_id = ws[row][0].value
    staff_name = ws[row][1].value
    staff_login = ws[row][2].value
    staff_role = ws[row][3].value

    # print(staff_name) # Вывести все данные столбца B (Фио)

    # Загоним в список значения каждой строки в листе
    staff_list = [staff_id, staff_name, staff_login, staff_role]
    print(staff_list)   # [11, 'Васечкин Петя', 'vasechkin', 'Тестировщик'] каждая строка в виде списка

    # Добавим в список списков все значения листа
    staff_row_list.append(staff_list)

print(staff_row_list)   # [['ID', 'Фио', 'Логин', 'Роль'], [11, 'Васечкин Петя', 'vasechkin', 'Тестировщик']]