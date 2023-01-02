# Общие данные по работе с Excel через openpyxl
# https://openpyxl.readthedocs.io/en/stable/usage.html

# Подключили openpyxl
import openpyxl
from openpyxl import *


# ----------------------------------------------------------------------------------------------------------------------
print('SECTION 1')
# Подключаем эксельку с нашей базой
wb = load_workbook("database.xlsx")      # wb -> wb
print(wb)                                # <openpyxl.workbook.workbook.Workbook object at 0x1059ce0d0>
print(type(wb))                          # ТИП: <class 'openpyxl.workbook.workbook.Workbook'>
print('-------------------------------------------')

# ----------------------------------------------------------------------------------------------------------------------
print('SECTION 2')
#  Получить активный лист книги
ws = wb.active          # ws -> ws
print(ws)               # <Worksheet "worklog">
print(type(ws))         # ТИП: <class 'openpyxl.worksheet.worksheet.Worksheet'>
print('-------------------------------------------')

# ----------------------------------------------------------------------------------------------------------------------
print('SECTION 3')
# Полный список всех листов книги в списке
wb_sheets_list = wb.sheetnames
print(wb_sheets_list)                         # ['staff', 'worklog']
print(wb_sheets_list[0])                      # Обратиться к 1му листу книги -> staff
print(type(wb_sheets_list))                   # ТИП: <class 'list'>
print(type(wb_sheets_list[0]))                # ТИП: <class 'str'>
print('-------------------------------------------')

# ----------------------------------------------------------------------------------------------------------------------
print('SECTION 4')
# Записываем листы в переменные
ws_staff = wb['staff']          # ws -> ws_staff
ws_worklog = wb['worklog']      # ws -> ws_worklog
print(ws_staff)                 # <Worksheet "staff">
print(type(ws_staff))           # ТИП: <class 'openpyxl.worksheet.worksheet.Worksheet'>
print('-------------------------------------------')

# ----------------------------------------------------------------------------------------------------------------------
print('SECTION 5')
# row       строки
# column    столбцы
# cell      ячейка
# Определяем количество строк на листе / staff
row_count_ws_staff = ws_staff.max_row
print(f'Строк: {row_count_ws_staff}')
# Определяем количество столбцов на листе / staff
column_count_ws_staff = ws_staff.max_column
print(f'Столбцов: {column_count_ws_staff}')
print('-------------------------------------------')

# ----------------------------------------------------------------------------------------------------------------------
print('SECTION 6')
# Определяем заголовки столбцов листа staff
rows = ws_staff.rows
print(rows)             # <generator object Worksheet._cells_by_row at 0x103cb5900>
print(type(rows))       # <class 'generator'>
# Пишем заголовки столбцов в список ws_headers
ws_headers = [cell.value for cell in next(rows)]
print(ws_headers)       # ['ID', 'Фио', 'Логин', 'Роль']
print('-------------------------------------------')
# ??? ws.append(ws_headers)

# ----------------------------------------------------------------------------------------------------------------------
print('SECTION 7')
# Обращение к ячейке
# Пример: на листе staff в ячейке B3 содержится значение 'Ершов Мирон'
# ws_staff[ROW][COLUMN] / Причем ROW начинается с 1! а COLUMN с 0!
# ws_staff[3][] -> третья строка на листе / ws_staff[][1] -> Столбец B, потому что столбец А это [0]
b3 = ws_staff[3][1].value
print(b3)
print('-------------------------------------------')

# ----------------------------------------------------------------------------------------------------------------------
print('SECTION 8')
# Выведем заголовки через обращения к ячейкам
a1 = ws_staff[1][0].value
b1 = ws_staff[1][1].value
c1 = ws_staff[1][2].value
d1 = ws_staff[1][3].value
print(a1,b1,c1,d1)
print('-------------------------------------------')

# ----------------------------------------------------------------------------------------------------------------------
print('SECTION 9 - Вариант 1')
# Прочитаем лист и загоним в список
# ws_staff.max_row - последний ряд (строку в таблице) не выведет, чтобы вывел надо {ws_staff.max_row+1}

# Создадим пустой список
staff_list = []
# Создадим пустой список списков
staff_list_of_list: list = []

for row in range(1, ws_staff.max_row+1): # (2, ws_staff.max_row+1) -> бе заголовка, начнет со второй строки
    # print(ws_staff[row])    # вывод: (<Cell 'staff'.A1>, <Cell 'staff'.B1>, <Cell 'staff'.C1>)
    staff_id = ws_staff[row][0].value
    staff_name = ws_staff[row][1].value
    staff_login = ws_staff[row][2].value
    staff_role = ws_staff[row][3].value

    # print(staff_name) # Получить все данные столбца B (Фио)

    # Загоним в список значения каждой строки в листе
    staff_list = [staff_id, staff_name, staff_login, staff_role]
    print(staff_list)   # [11, 'Васечкин Петя', 'vasechkin', 'Тестировщик'] каждая строка в виде списка

    # Добавим в список списков все значения листа
    staff_list_of_list.append(staff_list)

print(staff_list_of_list)   # [['ID', 'Фио', 'Логин', 'Роль'], [11, 'Васечкин Петя', 'vasechkin', 'Тестировщик']]
print('-------------------------------------------')

# ----------------------------------------------------------------------------------------------------------------------
print('SECTION 10 - Вариант 2')
# ПРЕДПОЧТИТЕЛЬНЫЙ МЕТОД ДЛЯ ПЕРЕБОРА ТАБЛИЦ
# Прочитаем лист и загоним в список
# min_row=1 -> с заголовком
# min_row=2 -> без заголовка

# Создали пустой список списков
staff_row_list: list = []

for row in ws_staff.iter_rows(min_row=1, max_row=ws_staff.max_row, min_col=1, max_col=ws_staff.max_column):
    row = [cell.value for cell in row]
    # Добавляем данные в список
    staff_row_list.append(row)

print(staff_row_list)
print('-------------------------------------------')

# ----------------------------------------------------------------------------------------------------------------------
print('SECTION 11 - Вариант 3')

# Постучаться к листу (staff) можно одним из 3х способов / <Worksheet "staff">
# sheet = wb.active         # по активному листу
# sheet = wb.worksheets[0]  #  по порядковому номеру в книге, если в файле поменять местами листы - то все поломается
sheet = wb['staff']         # по конкретному названию листа в книге

for row in sheet.values:
    for cell in row:
        print(cell)

# ИЛИ тоже самое

# for row in sheet:
#     for cell in row:
#         print(cell.value)
#         # print(cell.font) # получить значение шрифта
print('-------------------------------------------')

# ----------------------------------------------------------------------------------------------------------------------
print('SECTION 12 - ДОБАВИТЬ ЗАГОЛОВКИ СТОЛБЦОВ В ФАЙЛЕ')
sheet = wb['staff']         # по конкретному названию листа в книге

# worksheet.insert_rows(0)
# worksheet['A1'].value = 'issue_key'
# worksheet['B1'].value = 'issue_summary'
# worksheet['C1'].value = 'hours'
# worksheet['D1'].value = 'work_date'
# worksheet['E1'].value = 'user_login'